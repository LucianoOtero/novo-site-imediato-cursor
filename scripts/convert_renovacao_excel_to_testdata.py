# -*- coding: utf-8 -*-
"""Converte a planilha de renovação em base local de testes RPA (JSON + MD)."""

from __future__ import annotations

import json
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl

SRC = Path(
    r"C:\Users\Luciano\OneDrive - Imediato Soluções em Seguros\Imediato"
    r"\imediatoseguros-rpa-playwright\Renovação teste site Luciano.xlsx"
)
OUT_RPA_DIR = Path(
    r"C:\Users\Luciano\OneDrive - Imediato Soluções em Seguros\Imediato"
    r"\imediatoseguros-rpa-playwright\data\renovacao_teste_site_luciano"
)
OUT_SITE_DIR = Path(
    r"C:\Users\Luciano\OneDrive - Imediato Soluções em Seguros\Imediato"
    r"\Novo Site Imediato Cursor\testdata\rpa"
)

MOTO_BRANDS = {
    "YAMAHA",
    "HONDA",
    "KAWASAKI",
    "SUZUKI",
    "BMW",
    "BAJAJ",
    "TRIUMPH",
    "HAOJUE",
    "DAFRA",
    "SHINERAY",
    "ROYAL ENFIELD",
    "KTM",
    "DUCATI",
    "HARLEY",
    "HARLEY-DAVIDSON",
    "MV AGUSTA",
    "APRILIA",
    "BENELLI",
}
MOTO_KEYWORDS = re.compile(
    r"\b(CG|BIZ|PCX|NXR|BROS|CBX|TWISTER|XRE|ADV|ELITE|SH\s|NEO|CRYPTON|"
    r"YBR|FACTOR|DOMINAR|STREET TRIPLE|G\s*310|XVS|DRAG|NK\s|NH\s|Z-?800|"
    r"SCOOTER|MOTOCICLETA|MOTO)\b",
    re.I,
)
CAMINHAO_KEYWORDS = re.compile(r"\b(CARGO|DAILY|TRUCK|CAMINH)\b", re.I)


def only_digits(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\D", "", str(v))


def pad_cpf_cnpj(digits: str) -> str:
    if not digits:
        return ""
    if len(digits) <= 11:
        return digits.zfill(11)
    return digits.zfill(14)


def pad_cep(digits) -> str:
    d = only_digits(digits)
    return d.zfill(8) if d else ""


def format_cep(digits: str) -> str:
    d = pad_cep(digits)
    return f"{d[:5]}-{d[5:]}" if len(d) == 8 else d


def format_date(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")
    s = str(v).strip()
    return s or None


def iso_date(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v)


def parse_phone(v) -> tuple[str, str, str]:
    d = only_digits(v)
    if not d:
        return "", "", ""
    if len(d) >= 11:
        return d[:2], d[2:], d
    if len(d) == 10:
        return d[:2], "9" + d[2:], d[:2] + "9" + d[2:]
    if len(d) == 9:
        return "", d, d
    return "", d, d


def normalize_sexo(v) -> str:
    if not v:
        return ""
    s = str(v).strip().lower()
    if s.startswith("m"):
        return "Masculino"
    if s.startswith("f"):
        return "Feminino"
    return str(v).strip()


def normalize_estado_civil(v) -> str:
    if not v:
        return ""
    s = str(v).strip()
    mapping = {
        "casado": "Casado",
        "solteiro": "Solteiro",
        "divorciado": "Divorciado",
        "viuvo": "Viuvo",
        "viúvo": "Viuvo",
        "uniao estavel": "Casado ou Uniao Estavel",
        "união estável": "Casado ou Uniao Estavel",
    }
    return mapping.get(s.lower(), s)


def detect_tipo_veiculo(marca: str, modelo: str) -> str:
    marca_u = (marca or "").upper().strip()
    modelo_s = modelo or ""
    marca_key = marca_u.split("-")[0].strip() if marca_u else ""
    if CAMINHAO_KEYWORDS.search(modelo_s):
        return "caminhao"
    if marca_key in MOTO_BRANDS or MOTO_KEYWORDS.search(modelo_s):
        if marca_key == "BMW" and not MOTO_KEYWORDS.search(modelo_s):
            return "carro"
        if marca_key == "HONDA" and not MOTO_KEYWORDS.search(modelo_s):
            if re.search(r"\b(CIVIC|CITY|FIT|HR-?V|CR-?V|WR-?V|ACCORD)\b", modelo_s, re.I):
                return "carro"
            return "moto"
        return "moto"
    return "carro"


def infer_ramo(tipo_veiculo: str) -> str:
    if tipo_veiculo == "moto":
        return "moto"
    if tipo_veiculo == "caminhao":
        return "caminhao"
    return "auto"


def slugify(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s.lower()).strip("-")
    return s[:48] or "caso"


def main() -> None:
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    cases = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not any(row):
            continue
        data = dict(zip(headers, row))
        nome = str(data.get("Segurado") or "").strip()
        doc = pad_cpf_cnpj(only_digits(data.get("CNPJ\\CPF Segurado")))
        placa = str(data.get("Placa") or "").strip().upper().replace("-", "").replace(" ", "")
        cep_digits = pad_cep(data.get("CEP"))
        ddd, celular, celular_full = parse_phone(
            data.get("Fone Celular") or data.get("Fone Residencial")
        )
        marca = str(data.get("Marca") or "").strip()
        modelo = str(data.get("Modelo") or "").strip()
        ano_fab = data.get("Ano Auto")
        ano_mod = data.get("Ano Modelo")
        is_pj = len(doc) > 11
        tipo_veiculo = detect_tipo_veiculo(marca, modelo)
        ramo = infer_ramo(tipo_veiculo)
        zero_km_val = row[17]
        zero_km = str(zero_km_val or "").strip().lower() in {"sim", "s", "yes", "true", "1"}

        case_id = f"ren-{idx:03d}-{slugify(nome)}"
        tags = [ramo, tipo_veiculo, "pj" if is_pj else "pf"]
        if row[38]:
            tags.append("cancelado")
        if not placa:
            tags.append("sem-placa")
        if not celular:
            tags.append("sem-celular")

        rpa_start_payload = {
            "DDD-CELULAR": f"{ddd}-{celular}" if ddd and celular else "",
            "CELULAR": celular,
            "NOME": nome,
            "CPF": doc if not is_pj else "",
            "CNPJ": doc if is_pj else "",
            "CEP": cep_digits,
            "PLACA": placa,
            "produto": ramo,
        }

        rpa_parametros = {
            "tipo_veiculo": "moto" if tipo_veiculo == "moto" else "carro",
            "placa": placa,
            "marca": marca,
            "modelo": modelo,
            "ano": str(ano_mod or ano_fab or ""),
            "ano_fabricacao": str(ano_fab or ""),
            "ano_modelo": str(ano_mod or ""),
            "zero_km": zero_km,
            "cep": format_cep(cep_digits),
            "endereco_completo": f"CEP {format_cep(cep_digits)}" if cep_digits else "",
            "uso_veiculo": "Pessoal",
            "nome": nome,
            "cpf": doc if not is_pj else "",
            "cnpj": doc if is_pj else "",
            "data_nascimento": format_date(data.get("Data de Nascimento")),
            "sexo": normalize_sexo(data.get("Sexo")),
            "estado_civil": normalize_estado_civil(data.get("Estado Civil")),
            "email": str(data.get("Email") or "").strip(),
            "celular": celular_full or (ddd + celular if ddd else celular),
            "endereco": f"CEP {format_cep(cep_digits)}" if cep_digits else "",
            "condutor_principal": True,
            "logradouro": str(data.get("Logradouro") or "").strip(),
            "numero": str(row[28] or "").strip(),
            "bairro": str(data.get("Bairro") or "").strip(),
            "complemento": str(data.get("Complemento") or "").strip() if data.get("Complemento") else "",
            "cidade": str(data.get("Cidade") or "").strip(),
            "uf": str(data.get("UF") or "").strip(),
            "chassis": str(data.get("Chassis") or "").strip(),
            "referencia_fipe": str(row[19] or ""),
            "classe_bonus": row[39],
            "apolice": str(row[1] or ""),
            "fim_vigencia": format_date(row[5]),
            "continuar_com_corretor_anterior": True,
        }

        form_fields = {
            "nome": nome,
            "cpf": doc if not is_pj else "",
            "cnpj": doc if is_pj else "",
            "ddd": ddd,
            "celular": celular,
            "email": str(data.get("Email") or "").strip(),
            "cep": cep_digits,
            "placa": placa,
            "ramo": ramo,
            "veiculoMarca": marca,
            "veiculoModelo": modelo,
            "veiculoAnoFabricacao": str(ano_fab or ""),
            "veiculoAnoModelo": str(ano_mod or ""),
            "dataNascimento": format_date(data.get("Data de Nascimento")),
            "sexo": normalize_sexo(data.get("Sexo")),
            "estadoCivil": normalize_estado_civil(data.get("Estado Civil")),
            "cidade": str(data.get("Cidade") or "").strip(),
            "uf": str(data.get("UF") or "").strip(),
        }

        reasons = []
        if is_pj:
            reasons.append("pessoa_juridica")
        if not placa:
            reasons.append("sem_placa")
        if not ddd or not celular:
            reasons.append("sem_celular_valido")
        if not cep_digits:
            reasons.append("sem_cep")
        if not nome:
            reasons.append("sem_nome")

        case = {
            "id": case_id,
            "index": idx,
            "tags": tags,
            "elegivel_rpa_site": len(reasons) == 0,
            "motivo_inelegivel": reasons or None,
            "segurado": {
                "nome": nome,
                "documento": doc,
                "tipoPessoa": "Juridica" if is_pj else "Fisica",
                "dataNascimento": format_date(data.get("Data de Nascimento")),
                "sexo": normalize_sexo(data.get("Sexo")),
                "estadoCivil": normalize_estado_civil(data.get("Estado Civil")),
                "email": str(data.get("Email") or "").strip(),
                "telefoneResidencial": str(data.get("Fone Residencial") or "") or None,
                "telefoneCelular": str(data.get("Fone Celular") or "") or None,
            },
            "endereco": {
                "cep": cep_digits,
                "cepFormatado": format_cep(cep_digits),
                "logradouro": str(data.get("Logradouro") or "").strip(),
                "numero": str(row[28] or "").strip(),
                "bairro": str(data.get("Bairro") or "").strip(),
                "complemento": str(data.get("Complemento") or "").strip() if data.get("Complemento") else None,
                "cidade": str(data.get("Cidade") or "").strip(),
                "uf": str(data.get("UF") or "").strip(),
            },
            "veiculo": {
                "marca": marca,
                "modelo": modelo,
                "anoFabricacao": str(ano_fab or ""),
                "anoModelo": str(ano_mod or ""),
                "placa": placa,
                "chassis": str(data.get("Chassis") or "").strip(),
                "zeroKm": zero_km,
                "tipoVeiculo": tipo_veiculo,
                "referenciaFipe": str(row[19] or ""),
                "fatorAjuste": data.get("Fator de Ajuste"),
            },
            "apolice": {
                "numero": str(row[1] or ""),
                "inicioVigencia": iso_date(row[4]),
                "fimVigencia": iso_date(row[5]),
                "classeBonus": row[39],
                "cancelamento": iso_date(row[38]) if row[38] else None,
                "qtdeParcelas": row[2],
                "valorTarifario": row[9],
                "valorIS": row[7],
            },
            "formFields": form_fields,
            "rpaStartPayload": rpa_start_payload,
            "rpaParametros": rpa_parametros,
        }
        cases.append(case)

    OUT_RPA_DIR.mkdir(parents=True, exist_ok=True)
    OUT_SITE_DIR.mkdir(parents=True, exist_ok=True)

    meta = {
        "fonte": "Renovação teste site Luciano.xlsx",
        "geradoEm": datetime.now().isoformat(timespec="seconds"),
        "totalCasos": len(cases),
        "elegiveisRpaSite": sum(1 for c in cases if c["elegivel_rpa_site"]),
        "porRamo": {},
        "porTipoVeiculo": {},
        "notas": [
            "rpaStartPayload segue o contrato de lib/rpa.ts (buildRpaPayload) do site novo.",
            "rpaParametros segue o formato usado em testes Playwright do motor RPA (sem autenticacao).",
            "Casos PJ ficam marcados como inelegiveis para o fluxo PF do site.",
            "tipo_veiculo/ramo sao inferidos por marca/modelo — revisar casos duvidosos antes de regressao.",
            "Contem PII real de clientes — NAO versionar em repositorio publico.",
        ],
    }
    for c in cases:
        meta["porRamo"][c["formFields"]["ramo"]] = meta["porRamo"].get(c["formFields"]["ramo"], 0) + 1
        tv = c["veiculo"]["tipoVeiculo"]
        meta["porTipoVeiculo"][tv] = meta["porTipoVeiculo"].get(tv, 0) + 1

    db = {"meta": meta, "casos": cases}
    payload = json.dumps(db, ensure_ascii=False, indent=2)

    (OUT_RPA_DIR / "casos.json").write_text(payload, encoding="utf-8")
    (OUT_SITE_DIR / "renovacao-teste-site-luciano.json").write_text(payload, encoding="utf-8")

    slim = {
        "meta": {**meta, "subset": "elegiveis_rpa_site", "totalCasos": meta["elegiveisRpaSite"]},
        "casos": [
            {
                "id": c["id"],
                "tags": c["tags"],
                "formFields": c["formFields"],
                "rpaStartPayload": c["rpaStartPayload"],
            }
            for c in cases
            if c["elegivel_rpa_site"]
        ],
    }
    slim_payload = json.dumps(slim, ensure_ascii=False, indent=2)
    (OUT_SITE_DIR / "renovacao-teste-site-luciano.elegiveis.json").write_text(slim_payload, encoding="utf-8")
    (OUT_RPA_DIR / "casos.elegiveis.json").write_text(slim_payload, encoding="utf-8")

    lines = [
        "# Base de testes RPA — Renovação (Luciano)",
        "",
        f"**Fonte:** `{meta['fonte']}`  ",
        f"**Gerado em:** {meta['geradoEm']}  ",
        f"**Total:** {meta['totalCasos']} casos · **Elegíveis site RPA:** {meta['elegiveisRpaSite']}",
        "",
        "## Distribuição",
        "",
        "| Dimensão | Valor | Qtde |",
        "|---|---|---:|",
    ]
    for k, v in sorted(meta["porRamo"].items()):
        lines.append(f"| ramo | `{k}` | {v} |")
    for k, v in sorted(meta["porTipoVeiculo"].items()):
        lines.append(f"| tipoVeiculo | `{k}` | {v} |")
    lines += [
        "",
        "## Como usar",
        "",
        "### Site novo (`buildRpaPayload`)",
        "",
        "```ts",
        'import db from "@/testdata/rpa/renovacao-teste-site-luciano.elegiveis.json";',
        "for (const caso of db.casos) {",
        "  await startRpaSession(caso.rpaStartPayload);",
        "}",
        "```",
        "",
        "### Motor Playwright RPA",
        "",
        "Use `casos[i].rpaParametros` mesclado com `configuracao`/`autenticacao` locais.",
        "",
        "Arquivos:",
        "",
        "- `casos.json` / `renovacao-teste-site-luciano.json` — base completa",
        "- `casos.elegiveis.json` / `renovacao-teste-site-luciano.elegiveis.json` — só PF com placa/CEP/celular",
        "",
        "## Índice dos casos",
        "",
        "| # | id | nome | placa | ramo | CEP | celular | elegível |",
        "|---:|---|---|---|---|---|---|---|",
    ]
    for c in cases:
        el = "sim" if c["elegivel_rpa_site"] else "não"
        cel = c["formFields"]["ddd"] + c["formFields"]["celular"]
        lines.append(
            f"| {c['index']} | `{c['id']}` | {c['segurado']['nome']} | {c['veiculo']['placa']} | "
            f"{c['formFields']['ramo']} | {c['endereco']['cepFormatado']} | {cel} | {el} |"
        )
    lines += ["", "## Avisos", ""]
    for n in meta["notas"]:
        lines.append(f"- {n}")
    lines.append("")

    md = "\n".join(lines)
    (OUT_RPA_DIR / "README.md").write_text(md, encoding="utf-8")
    (OUT_SITE_DIR / "README.md").write_text(md, encoding="utf-8")

    print("OK")
    print("casos", len(cases))
    print("elegiveis", meta["elegiveisRpaSite"])
    print("porRamo", meta["porRamo"])
    print("porTipo", meta["porTipoVeiculo"])
    print("sample0", json.dumps(cases[0]["rpaStartPayload"], ensure_ascii=False))
    ine = [c for c in cases if not c["elegivel_rpa_site"]]
    print("inelegiveis", len(ine))
    for c in ine[:15]:
        print(" -", c["id"], c["motivo_inelegivel"])
    print("wrote", OUT_RPA_DIR)
    print("wrote", OUT_SITE_DIR)


if __name__ == "__main__":
    main()
